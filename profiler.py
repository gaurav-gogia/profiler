# echo 3 > /proc/sys/vm/drop_caches

import psutil
import subprocess
import sys
import time
import os
import threading
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

EXCEL_FILE = "profiler_results.xlsx"

HEADERS = [
    "Timestamp",
    "Run Label",
    "Sample Rate (s)",
    "Program",
    "Program Args",
    "Execution Time (s)",
    "Peak CPU (%)",
    "Avg CPU (%)",
    "Peak RAM (bytes)",
    "Avg RAM (bytes)",
    "Peak Threads",
    "Avg Threads",
    "Peak IO Read (bytes)",
    "Avg IO Read (bytes)",
    "Peak IO Write (bytes)",
    "Avg IO Write (bytes)",
    "Original File Size",
    "Size in DB",
]


def convert_bytes(num) -> str:
    for x in ["bytes", "KB", "MB", "GB", "TB"]:
        if num < 1024.0:
            return "%3.2f %s" % (num, x)
        num /= 1024.0


def save_to_excel(row_data: list):
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profiler Results"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    ws.append(row_data)
    wb.save(EXCEL_FILE)
    print(f"Results saved to {os.path.abspath(EXCEL_FILE)}")


def get_next_run_label() -> int:
    if not os.path.exists(EXCEL_FILE):
        return 1

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
    except Exception:
        return 1

    if ws is None:
        return 1

    # If only headers exist (or sheet is empty), start at 1.
    if ws.max_row <= 1:
        return 1

    previous_value = ws.cell(row=ws.max_row, column=2).value
    try:
        return int(float(str(previous_value).strip())) + 1
    except (TypeError, ValueError):
        return 1


def monitor_process(proc, sample_rate, run_label, program_name, program_args):
    p = psutil.Process(proc.pid)

    cpu_use_list = []
    mem_use_list = []
    thread_count_list = []
    read_bytes_list = []
    write_bytes_list = []

    # First cpu_percent() call always returns 0.0 — warm it up and discard
    p.cpu_percent()

    stop_event = threading.Event()

    def collect():
        while not stop_event.is_set():
            try:
                with p.oneshot():
                    cpu = p.cpu_percent() / os.cpu_count()
                    mem = p.memory_info().rss
                    threads = p.num_threads()
                    io = p.io_counters()
                cpu_use_list.append(cpu)
                mem_use_list.append(mem)
                thread_count_list.append(threads)
                read_bytes_list.append(io.read_bytes)
                write_bytes_list.append(io.write_bytes)
            except psutil.NoSuchProcess:
                break
            # stop_event.wait respects the event immediately when set,
            # avoiding a full extra sleep at the end
            stop_event.wait(sample_rate)

    start_time = time.perf_counter()
    monitor_thread = threading.Thread(target=collect, daemon=True)
    monitor_thread.start()

    proc.wait()  # OS-notified — no polling, precise termination detection
    end_time = time.perf_counter()

    stop_event.set()
    monitor_thread.join()

    execution_time = end_time - start_time
    print("\n\n")

    print(f"Execution time: {execution_time}")

    avg_cpu_use = sum(cpu_use_list) / len(cpu_use_list)
    peak_ram = max(mem_use_list)
    avg_ram = int(sum(mem_use_list) / len(mem_use_list))
    avg_thread_count = int(sum(thread_count_list) / len(thread_count_list))
    peak_io_read = max(read_bytes_list)
    avg_io_read = int(sum(read_bytes_list) / len(read_bytes_list))
    peak_io_write = max(write_bytes_list)
    avg_io_write = int(sum(write_bytes_list) / len(write_bytes_list))

    print(f"Peak CPU Use: {max(cpu_use_list)}% | Avg CPU Use: {avg_cpu_use}%")
    print(
        f"Peak RAM Use: {convert_bytes(peak_ram)} | Avg RAM Use: {convert_bytes(avg_ram)}"
    )
    print(f"Peak Threads: {max(thread_count_list)} | Avg Threads: {avg_thread_count}")
    print(
        f"Peak IO Read: {convert_bytes(peak_io_read)} | Avg IO Read: {convert_bytes(avg_io_read)}"
    )
    print(
        f"Peak IO Write: {convert_bytes(peak_io_write)} | Avg IO Write: {convert_bytes(avg_io_write)}"
    )
    print(f"Run Count: {run_label}")

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        run_label,
        sample_rate,
        program_name,
        " ".join(program_args),
        round(execution_time, 4),
        round(max(cpu_use_list), 4),
        round(avg_cpu_use, 4),
        peak_ram,
        avg_ram,
        max(thread_count_list),
        avg_thread_count,
        peak_io_read,
        avg_io_read,
        peak_io_write,
        avg_io_write,
        "",  # Original File Size — fill manually
        "",  # Size in DB — fill manually
    ]
    save_to_excel(row)


def main():
    if len(sys.argv) < 3:
        print("Usage: profiler.py <sample_rate> <program> [program_args...]")
        print("  sample_rate  Sampling interval in seconds (e.g. 0.1)")
        print("  program      Program to profile")
        print("  program_args Optional arguments passed to the program")
        sys.exit(1)

    sample_rate = float(sys.argv[1])
    run_label = get_next_run_label()
    program_name = sys.argv[2]
    program_args = sys.argv[3:]

    proc = subprocess.Popen([program_name] + program_args)
    monitor_process(proc, sample_rate, run_label, program_name, program_args)
    print("-----------------------------------------------------------------------------\n\n")


if __name__ == "__main__":
    main()
